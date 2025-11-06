import pandas as pd
import requests
from bs4 import BeautifulSoup as bs
import os
import glob

# Delete old result Excel files (keep only the CSV)
print("Cleaning up old result files...")
excel_files = glob.glob("chartink_*.xlsx")
for file in excel_files:
    try:
        os.remove(file)
        print(f"   Deleted: {file}")
    except:
        pass
print("   [OK] Cleanup complete\n")

# API endpoint for processing the screener
api_url = "https://chartink.com/screener/process"

# URLs for gainers (Open = High) and losers (Open = Low)
gainers_url = "https://chartink.com/screener/copy-open-high-5911"
losers_url = "https://chartink.com/screener/copy-open-low-103152"

def fetch_stocks(session, screener_url, condition_type="high"):
    """Fetch stocks from Chartink API"""
    print(f"\n{'='*60}")
    print(f"Fetching {condition_type.upper()} data from Chartink...")
    print(f"{'='*60}")
    
    # Get CSRF token
    print(f"\nStep 1: Getting CSRF token from {condition_type} screener...")
    r_data = session.get(screener_url)
    soup = bs(r_data.content, "lxml")
    meta = soup.find("meta", {"name": "csrf-token"})["content"]
    print(f"   CSRF token: {meta[:20]}...")
    
    # Determine scan clause based on condition type
    if condition_type == "high":
        condition = {"scan_clause": "( latest open = latest high )"}
    else:  # low
        condition = {"scan_clause": "( latest open = latest low )"}
    
    print(f"\nStep 2: Fetching stock data...")
    print(f"   Condition: {condition['scan_clause']}")
    
    header = {"x-csrf-token": meta}
    response = session.post(api_url, headers=header, data=condition)
    
    if response.status_code == 200:
        data = response.json()
        if "data" in data and len(data["data"]) > 0:
            stock_list = pd.DataFrame(data["data"])
            print(f"   [OK] Successfully fetched {len(stock_list)} stocks")
            return stock_list
        else:
            print(f"   [WARNING] No data found")
            return pd.DataFrame()
    else:
        print(f"   [ERROR] API returned status code {response.status_code}")
        return pd.DataFrame()

def load_nifty100_list():
    """Load Nifty 100 stock symbols from CSV"""
    csv_file = "ind_nifty100list.csv"
    try:
        if os.path.exists(csv_file):
            nifty100_df = pd.read_csv(csv_file)
            # Extract symbols and normalize
            nifty100_symbols = nifty100_df['Symbol'].str.strip().str.upper().str.replace('-', '').tolist()
            nifty100_symbols_original = nifty100_df['Symbol'].str.strip().str.upper().tolist()
            nifty100_symbols = list(set(nifty100_symbols + nifty100_symbols_original))
            nifty100_symbols = [s for s in nifty100_symbols if s and s != 'NAN' and not pd.isna(s)]
            
            # Add ICICIPRULI if missing
            if 'ICICIPRULI' not in nifty100_symbols:
                nifty100_symbols.append('ICICIPRULI')
            
            return nifty100_symbols
        else:
            print(f"[WARNING] CSV file '{csv_file}' not found")
            return []
    except Exception as e:
        print(f"[ERROR] Failed to load CSV: {e}")
        return []

def filter_and_sort_stocks(stock_list, nifty100_symbols, condition_type="gainers"):
    """Filter to Nifty 100 and sort by percentage change"""
    if stock_list.empty:
        return stock_list
    
    # Find percentage change column
    pct_col = None
    for col in stock_list.columns:
        col_lower = str(col).lower()
        if 'chg' in col_lower or 'change' in col_lower or 'pct' in col_lower or '%' in col_lower:
            pct_col = col
            break
    
    # Filter to Nifty 100
    if 'nsecode' in stock_list.columns:
        nifty100_stocks = stock_list[stock_list['nsecode'].isin(nifty100_symbols)]
        print(f"\n[INFO] Filtered to {len(nifty100_stocks)} Nifty 100 stocks")
        stock_list = nifty100_stocks
    
    # Sort by percentage change
    if pct_col:
        stock_list = stock_list.copy()  # Avoid SettingWithCopyWarning
        stock_list[pct_col] = pd.to_numeric(stock_list[pct_col], errors='coerce')
        # For gainers: sort descending (best first)
        # For losers: sort descending (highest % first, like Chartink - ASIANPAINT 4.67% then AXISBANK 0.15%)
        ascending = False  # Both sorted descending by % change
        stock_list = stock_list.sort_values(pct_col, ascending=ascending, na_position='last')
        print(f"   [OK] Sorted by % Change ({'Best to Worst' if condition_type == 'gainers' else 'Highest to Lowest'})")
    
    return stock_list

# Main execution
print("\n" + "="*60)
print("NIFTY 100 - TOP GAINERS & TOP LOSERS")
print("="*60)

# Load Nifty 100 list
nifty100_symbols = load_nifty100_list()
if not nifty100_symbols:
    print("[ERROR] Could not load Nifty 100 list. Exiting.")
    exit(1)

print(f"\n[INFO] Loaded {len(nifty100_symbols)} Nifty 100 stocks from CSV")

# Create session
with requests.session() as s:
    # Fetch gainers (Open = High)
    gainers_df = fetch_stocks(s, gainers_url, "high")
    gainers_df = filter_and_sort_stocks(gainers_df, nifty100_symbols, "gainers")
    
    # Fetch losers (Open = Low)
    losers_df = fetch_stocks(s, losers_url, "low")
    losers_df = filter_and_sort_stocks(losers_df, nifty100_symbols, "losers")
    
    # Display results
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    print(f"\n[GAINERS] Top Gainers (Open = High): {len(gainers_df)} stocks")
    if not gainers_df.empty:
        print(gainers_df[['nsecode', 'name', 'per_chg', 'close', 'volume']].head(10).to_string(index=False))
    
    print(f"\n[LOSERS] Top Losers (Open = Low): {len(losers_df)} stocks")
    if not losers_df.empty:
        print(losers_df[['nsecode', 'name', 'per_chg', 'close', 'volume']].head(10).to_string(index=False))
    
    # Prepare combined data with proper formatting
    output_file = "nifty100_gainers_losers.xlsx"
    print(f"\n{'='*60}")
    print(f"Saving to Excel: {output_file}")
    print(f"{'='*60}")
    
    # Create combined dataframe with proper structure
    combined_data = []
    
    # Add gainers section
    if not gainers_df.empty:
        gainers_df_copy = gainers_df.copy()
        gainers_df_copy['Type'] = 'Gainer'
        gainers_df_copy['Section'] = 'Top Gainers (Open = High)'
        combined_data.append(gainers_df_copy)
    
    # Add losers section (already sorted by % change descending)
    if not losers_df.empty:
        losers_df_copy = losers_df.copy()
        losers_df_copy['Type'] = 'Loser'
        losers_df_copy['Section'] = 'Top Losers (Open = Low)'
        combined_data.append(losers_df_copy)
    
    if combined_data:
        # Combine both sections
        final_df = pd.concat(combined_data, ignore_index=True)
        
        # Reorder columns for better presentation
        column_order = ['Section', 'Type', 'sr', 'nsecode', 'name', 'bsecode', 'per_chg', 'close', 'volume']
        # Only include columns that exist
        available_cols = [col for col in column_order if col in final_df.columns]
        final_df = final_df[available_cols]
        
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Save to Excel
            final_df.to_excel(output_file, sheet_name='Gainers & Losers', index=False)
            
            # Apply formatting
            wb = load_workbook(output_file)
            ws = wb['Gainers & Losers']
            
            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Format data rows and add borders
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add section headers formatting (if Section column exists)
            if 'Section' in final_df.columns:
                section_col_idx = available_cols.index('Section') + 1
                current_section = None
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                    section_value = ws.cell(row=row_idx, column=section_col_idx).value
                    if section_value != current_section:
                        current_section = section_value
                        # Format section header row
                        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                        section_font = Font(bold=True, size=10)
                        for cell in ws[row_idx]:
                            cell.fill = section_fill
                            cell.font = section_font
            
            wb.save(output_file)
            
            print(f"\n[SUCCESS] Data saved to {output_file}")
            print(f"   - Top Gainers: {len(gainers_df)} stocks")
            print(f"   - Top Losers: {len(losers_df)} stocks")
            print(f"   - Total: {len(final_df)} stocks in one sheet")
            print(f"   - Formatting: Headers, borders, and auto-width applied")
        
        except PermissionError:
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file_timestamp = f"nifty100_gainers_losers_{timestamp}.xlsx"
            print(f"\n[WARNING] {output_file} is open in Excel")
            print(f"[OK] Saving to {output_file_timestamp} instead")
            final_df.to_excel(output_file_timestamp, sheet_name='Gainers & Losers', index=False)
        except Exception as e:
            print(f"\n[WARNING] Could not apply advanced formatting: {e}")
            print(f"   Saving with basic formatting...")
            final_df.to_excel(output_file, sheet_name='Gainers & Losers', index=False)
            print(f"   [OK] Saved to {output_file}")
    else:
        print(f"\n[WARNING] No data to save")

print(f"\n{'='*60}")
print("COMPLETE!")
print(f"{'='*60}")

